# This is a program to keep track of your flips in Runescape at the Grand Exchange.
# The main menu has two main features: Add item, Confirm merch
#
# Add item:
# 1. Enter the name of the item
# 2. Enter the price that you bought one for
# 3. Enter the price that you sold one for
# 4. The program presents you with possible profits for different amounts (on a successful flip)
# 5. Choosing to add the item asks for the amount and adds it. Canceling takes you back to the main menu
#
# Confirm merch:
# Pick a merch that you wish to mark as completed
# Press "OK" and the program highlights it green in excel.

import easygui
import datetime
from openpyxl import *
from openpyxl.styles import PatternFill
redFill = PatternFill(start_color='32cd32',
                   end_color='32cd32',
                   fill_type='solid')

# Displays possbile profit margins and adds the item if told so
def additem():
    item = easygui.enterbox("Add item","MerchGUI")
    item = item.upper()

    sellfor = int(easygui.enterbox("Bought for?","MerchGUI"))
    buyfor = int(easygui.enterbox("Sold for?","MerchGUI"))
    margin = sellfor - buyfor

    addit = easygui.buttonbox("Buy for: " + str(buyfor) + " / Sell for: " + str(sellfor) + "\n"
                              "Margin per item: " + str(margin) + "\n"
                      "Margin per 10: " + str(margin * 10) + "\n"
                      "Margin per 100: " + str(margin * 100) + "\n"
                      "Margin per 1k: " + str(margin * 1000) + "\n"
                      "Margin per 10k: " + str(margin * 10000) + "\n"
                      "Continue?","MerchGUI",choices=["Yes","No"],default_choice="Yes")
    if addit == "No":
        mainmenu()
    elif addit == "Yes":
        howmany = int(easygui.enterbox("How many?","MerchGUI"))
        totalgain = margin * howmany
        finalform = []
        finalform.append(str(item)) ##0
        finalform.append(str(buyfor)) ##1
        finalform.append(str(sellfor)) ##2
        finalform.append(str(howmany)) ##3
        finalform.append(str(totalgain)) ##4
        totable(finalform[0],finalform[1], finalform[2], finalform[3], finalform[4])
        easygui.buttonbox("Added " + finalform[3] + "x " + finalform[0] + "\n" + finalform[1] + "GP Buy\n" +
                       finalform[2] + "GP Sell\nTotal profit: " + finalform[4] + "GP!\nGood luck!","MerchGUI",
                          choices=["Ok"],default_choice="Ok")
        mainmenu()

# Appends it's arguments to the excel table.
def totable(item, buy, sell, qty, gain):

	# Loads the workbook and it's active sheet
    wb = load_workbook("Merch.xlsx")
    ws = wb.active

	# Adds information to first free row
    RowNum = ws.max_row + 1
    dateplace = "A" + str(RowNum)
    itemplace = "B" + str(RowNum)
    buyplace = "C" + str(RowNum)
    sellplace = "D" + str(RowNum)
    qtyplace = "E" + str(RowNum)
    gainplace = "F" + str(RowNum)
    ws[dateplace] = datetime.datetime.now()
    ws[itemplace] = item
    ws[buyplace] = int(buy)
    ws[sellplace] = int(sell)
    ws[qtyplace] = int(qty)
    ws[gainplace] = int(gain)

	# Saves workbook
    wb.save("Merch.xlsx")

#Confirms a merch
def confirm():

	# Loads the workbook and selects the active worksheet
    wb = load_workbook("Merch.xlsx")
    ws = wb.active

	# Adds table elements to choiclist which is displayed for user to select
    RowNum = ws.max_row
    i = 1
    choicelist = []
    while i < ws.max_row:
        dateplace = "A" + str(RowNum)
        itemplace = "B" + str(RowNum)
        buyplace = "C" + str(RowNum)
        sellplace = "D" + str(RowNum)
        qtyplace = "E" + str(RowNum)
        gainplace = "F" + str(RowNum)
        date = (str((ws[dateplace].value)).split("."))[0]
        qty = str(ws[qtyplace].value)
        item = str(ws[itemplace].value)
        gain = str(ws[gainplace].value)
        choicelist.append(str(i) + ". " + qty + "x " + item + " for " + gain + " on " + date)
        i += 1
        RowNum -= 1

	# Fills the chosen transaction
    completed = easygui.choicebox("Pick a trade and press Ok to mark it completed!","MerchGUI",choices=choicelist)
    completed = completed.split(".")
    RowNum = ws.max_row - int(completed[0])
    dateplace = "A" + str(RowNum)
    itemplace = "B" + str(RowNum)
    buyplace = "C" + str(RowNum)
    sellplace = "D" + str(RowNum)
    qtyplace = "E" + str(RowNum)
    gainplace = "F" + str(RowNum)
    ws[dateplace].fill = redFill
    ws[itemplace].fill = redFill
    ws[buyplace].fill = redFill
    ws[sellplace].fill = redFill
    ws[qtyplace].fill = redFill
    ws[gainplace].fill = redFill
    wb.save("Merch.xlsx")
    mainmenu()

# Main menu
def mainmenu():
    menuoption = easygui.buttonbox("Home","MerchGUI",choices=["Exit", "Add item", "Confirm merch"],default_choice="Add item")
    if menuoption == "Confirm merch":
        confirm()
    if menuoption == "Add item":
        additem()
    if menuoption == "Exit":
        exit()


mainmenu()